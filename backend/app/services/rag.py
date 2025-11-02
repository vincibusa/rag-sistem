from __future__ import annotations

import uuid
from collections.abc import Iterable
from dataclasses import asdict, is_dataclass
from io import BytesIO
from pathlib import Path
from tempfile import TemporaryDirectory
from typing import Any
import subprocess

import xlrd
from datapizza.type import Chunk
from openpyxl import load_workbook
from sqlalchemy.orm import Session

from app.core.config import settings
from app.core.exceptions import AppException
from app.core.logging import logger
from app.models import Document, DocumentChunk, DocumentStatus

from ..rag import (
    create_ingestion_pipeline,
    create_retrieval_pipeline,
    ensure_collection,
    get_vectorstore,
)
from ..rag.components import OllamaQueryEmbedder


class DocumentProcessingService:
    """Coordinate ingestion pipeline execution and persistence."""

    SUPPORTED_EXTENSIONS = {"pdf", "doc", "docx", "xls", "xlsx", "txt"}

    def __init__(self, session: Session):
        self.session = session

    def process_document(self, document_id: uuid.UUID) -> None:
        document = self.session.get(Document, document_id)
        if document is None:
            logger.warning("Documento %s non trovato: impossibile processarlo", document_id)
            return

        document.status = DocumentStatus.PROCESSING
        self.session.commit()

        try:
            self._process(document)
        except Exception as exc:
            logger.exception("Errore durante l'ingestione del documento %s", document_id)
            self.session.rollback()
            document = self.session.get(Document, document_id)
            if document:
                document.status = DocumentStatus.FAILED
                document.extra_metadata = {
                    **(document.extra_metadata or {}),
                    "last_error": str(exc),
                }
                self.session.commit()
            raise

    def _process(self, document: Document) -> None:
        ensure_collection()
        vectorstore = get_vectorstore()

        with TemporaryDirectory(prefix="rag-doc-") as tmp_dir:
            file_path, pipeline_kind = self._prepare_document_file(document, Path(tmp_dir))

            # Cleanup previous chunks if we are re-processing
            self._delete_existing_chunks(document, vectorstore)

            pipeline = create_ingestion_pipeline(pipeline_kind)
            metadata = {
                "document_id": str(document.id),
                "filename": document.filename,
                "content_type": document.content_type,
            }

            result = pipeline.run(file_path=str(file_path), metadata=metadata)
            chunks: list[Chunk] = list(result or [])

            if not chunks:
                raise RuntimeError("Nessun chunk generato dall'ingestione del documento.")

            enriched_chunks = self._enrich_chunks(chunks, document)
            logger.info(
                "Indicizzazione documento %s: %s chunk pronti per Qdrant",
                document.id,
                len(enriched_chunks),
            )
            vectorstore.add(enriched_chunks, collection_name=settings.qdrant_collection_name)
            logger.info(
                "Documento %s indicizzato correttamente nella collezione %s",
                document.id,
                settings.qdrant_collection_name,
            )
            self._persist_chunks(document, enriched_chunks)

        document.status = DocumentStatus.READY
        document.extra_metadata = {
            **(document.extra_metadata or {}),
            "embedding_model": settings.ollama_embed_model,
            "chunks_count": len(document.chunks),
        }
        self.session.commit()

    def _prepare_document_file(self, document: Document, base_dir: Path) -> tuple[Path, str]:
        extension = Path(document.filename).suffix.lower().lstrip(".")
        if extension not in self.SUPPORTED_EXTENSIONS:
            raise AppException(f"Formato {extension} non supportato per l'ingestione.")

        relative_path = Path(document.filename)
        if relative_path.is_absolute() or any(part == ".." for part in relative_path.parts):
            raise AppException("Percorso documento non valido nell'archivio caricato.")

        target_path = base_dir / relative_path
        target_path.parent.mkdir(parents=True, exist_ok=True)

        if extension == "doc":
            target_path.write_bytes(document.data)
            converted_path = self._convert_doc_to_docx(target_path)
            return converted_path, "docling"

        if extension in {"xls", "xlsx"}:
            text = self._extract_spreadsheet_text(document.data, extension)
            target_path = target_path.with_suffix(".txt")
            target_path.parent.mkdir(parents=True, exist_ok=True)
            target_path.write_text(text, encoding="utf-8")
            return target_path, "text"

        if extension == "txt":
            target_path.parent.mkdir(parents=True, exist_ok=True)
            target_path.write_bytes(document.data)
            return target_path, "text"

        target_path.write_bytes(document.data)
        return target_path, "docling"

    def _convert_doc_to_docx(self, source_path: Path) -> Path:
        binary = settings.libreoffice_binary or "soffice"
        output_dir = source_path.parent

        try:
            result = subprocess.run(
                [
                    binary,
                    "--headless",
                    "--convert-to",
                    "docx",
                    "--outdir",
                    str(output_dir),
                    str(source_path),
                ],
                check=True,
                stdout=subprocess.PIPE,
                stderr=subprocess.PIPE,
            )
        except FileNotFoundError as exc:
            raise AppException(
                "Impossibile convertire file .doc: comando LibreOffice 'soffice' non trovato. "
                "Installa LibreOffice e configura LIBREOFFICE_BINARY nel file .env.",
                status_code=500,
            ) from exc
        except subprocess.CalledProcessError as exc:  # pragma: no cover - dipendenza esterna
            logger.error(
                "Conversione LibreOffice fallita per %s: %s",
                source_path,
                exc.stderr.decode(errors="ignore"),
            )
            raise AppException(
                "Conversione del file .doc non riuscita. Verifica che il file non sia corrotto.",
                status_code=500,
            ) from exc

        logger.debug(
            "LibreOffice conversion output for %s: %s",
            source_path,
            result.stdout.decode(errors="ignore"),
        )

        converted_path = source_path.with_suffix(".docx")
        if not converted_path.exists():
            raise AppException(
                "Conversione .doc -> .docx non riuscita: file risultante mancante.",
                status_code=500,
            )
        return converted_path

    def _extract_spreadsheet_text(self, data: bytes, extension: str) -> str:
        lines: list[str] = []
        if extension == "xlsx":
            workbook = load_workbook(BytesIO(data), data_only=True, read_only=True)
            for sheet in workbook.worksheets:
                lines.append(f"# Foglio: {sheet.title}")
                for row in sheet.iter_rows(values_only=True):
                    values = [self._format_cell(value) for value in row]
                    if any(values):
                        lines.append(" | ".join(values))
        else:
            workbook = xlrd.open_workbook(file_contents=data)
            for sheet in workbook.sheets():
                lines.append(f"# Foglio: {sheet.name}")
                for row_idx in range(sheet.nrows):
                    row = sheet.row(row_idx)
                    values = [self._format_cell(cell.value) for cell in row]
                    if any(values):
                        lines.append(" | ".join(values))
        return "\n".join(lines)

    def _delete_existing_chunks(self, document: Document, vectorstore) -> None:
        point_ids = [
            chunk.qdrant_point_id for chunk in document.chunks if chunk.qdrant_point_id
        ]
        if point_ids:
            vectorstore.remove(
                collection_name=settings.qdrant_collection_name,
                ids=point_ids,
            )
        for chunk in list(document.chunks):
            self.session.delete(chunk)
        self.session.flush()

    def _enrich_chunks(self, chunks: Iterable[Chunk], document: Document) -> list[Chunk]:
        enriched: list[Chunk] = []
        for index, chunk in enumerate(chunks):
            embeddings = getattr(chunk, "embeddings", []) or []
            if not embeddings:
                raise RuntimeError(
                    "Il chunk generato non contiene embedding; verifica l'integrazione con Ollama."
                )
            vector = getattr(embeddings[0], "vector", None)
            if vector is None or len(vector) != settings.rag_embedding_dimensions:
                raise RuntimeError(
                    "Dimensione embedding inattesa per il chunk elaborato dal documento."
                )
            if not getattr(chunk, "id", None):
                chunk.id = str(uuid.uuid4())
            metadata = dict(getattr(chunk, "metadata", {}) or {})
            metadata.update(
                {
                    "document_id": str(document.id),
                    "chunk_index": index,
                    "source": document.filename,
                    "document_name": document.filename,
                }
            )
            chunk.metadata = metadata
            enriched.append(chunk)
        return enriched

    def _persist_chunks(self, document: Document, chunks: Iterable[Chunk]) -> None:
        for idx, chunk in enumerate(chunks):
            extra = self._to_serialisable_metadata(chunk.metadata)
            doc_chunk = DocumentChunk(
                document_id=document.id,
                chunk_index=idx,
                content=chunk.text,
                token_count=self._estimate_tokens(chunk.text),
                qdrant_point_id=getattr(chunk, "id", None),
                extra_metadata=extra,
            )
            document.chunks.append(doc_chunk)
        self.session.flush()

    @staticmethod
    def _to_serialisable_metadata(metadata: Any) -> dict[str, Any]:
        if metadata is None:
            return {}
        if is_dataclass(metadata):
            return asdict(metadata)
        if isinstance(metadata, dict):
            return metadata
        return {"value": str(metadata)}

    @staticmethod
    def _estimate_tokens(text: str) -> int:
        return max(1, len(text.split()))

    @staticmethod
    def _format_cell(value: Any) -> str:
        if value is None:
            return ""
        if isinstance(value, float):
            if value.is_integer():
                return str(int(value))
            return f"{value:.2f}"
        return str(value).strip()


class RagRetrievalService:
    """Execute the retrieval pipeline and return the generated answer with diagnostics."""

    def __init__(self) -> None:
        ensure_collection()
        self.vectorstore = get_vectorstore()
        self._query_embedder = OllamaQueryEmbedder()
        self._vector_name = settings.rag_embedding_name

    def run(self, query: str, *, top_k: int | None = None) -> dict[str, Any]:
        pipeline = create_retrieval_pipeline()
        k_value = top_k or settings.rag_top_k
        result = pipeline.run(
            {
                "rewriter": {"user_prompt": query},
                "prompt": {"user_prompt": query},
                "retriever": {
                    "collection_name": settings.qdrant_collection_name,
                    "k": k_value,
                    "vector_name": self._vector_name,
                },
                "generator": {
                    "input": query,
                    "system_prompt": None,
                },
            }
        )

        rewritten = _extract_text(result.get("rewriter"))
        answer = _extract_text(result.get("generator"))
        retrieved_chunks = result.get("retriever") or []

        chunks_payload = []
        for chunk in retrieved_chunks:
            metadata = self._to_serialisable(chunk.metadata)
            entry: dict[str, Any] = {
                "id": getattr(chunk, "id", None),
                "text": getattr(chunk, "text", ""),
                "metadata": metadata,
                "document_id": metadata.get("document_id"),
                "document_name": metadata.get("document_name") or metadata.get("source"),
                "chunk_index": metadata.get("chunk_index"),
            }
            score = getattr(chunk, "score", None)
            if score is not None:
                entry["score"] = score
            chunks_payload.append(entry)

        return {
            "query": query,
            "rewritten_query": rewritten,
            "answer": answer,
            "chunks": chunks_payload,
        }

    @staticmethod
    def _to_serialisable(metadata: Any) -> dict[str, Any]:
        if metadata is None:
            return {}
        if isinstance(metadata, dict):
            return metadata
        if is_dataclass(metadata):
            return asdict(metadata)
        return {"value": str(metadata)}

    def semantic_search(self, query: str, *, top_k: int | None = None) -> list[Chunk]:
        vector = self._query_embedder.embed_text(query)
        results = self.vectorstore.search(
            collection_name=settings.qdrant_collection_name,
            query_vector=vector,
            k=top_k or settings.rag_top_k,
            vector_name=self._vector_name,
        )
        return list(results)


def _extract_text(value: Any) -> str | None:
    if value is None:
        return None
    if hasattr(value, "text"):
        return getattr(value, "text")
    if isinstance(value, str):
        return value
    if isinstance(value, dict) and "text" in value:
        return value["text"]
    return str(value)
